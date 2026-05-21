from __future__ import annotations

import hashlib
import re
import uuid
from io import BytesIO
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.dbops_assets import (
    BusinessSystem,
    BusinessSystemContact,
    Cluster,
    ClusterVip,
    Contact,
    DbInstance,
    DbType,
    DbVersion,
    OsVersion,
    Server,
    Site,
    StagingExcelImport,
    TopologyRelation,
)


DB_TYPE_CATALOG = {
    "Oracle": {
        "type_code": "ORACLE",
        "name": "Oracle",
        "category": "relational",
        "license_type": "commercial",
        "vendor": "Oracle",
        "remark": "Oracle Database",
    },
    "PostgreSQL": {
        "type_code": "POSTGRESQL",
        "name": "PostgreSQL",
        "category": "relational",
        "license_type": "open_source",
        "vendor": "PostgreSQL",
        "remark": "Advanced open source RDBMS",
    },
    "SQL Server": {
        "type_code": "SQLSERVER",
        "name": "SQL Server",
        "category": "relational",
        "license_type": "commercial",
        "vendor": "Microsoft",
        "remark": "Microsoft SQL Server",
    },
    "MySQL": {
        "type_code": "MYSQL",
        "name": "MySQL",
        "category": "relational",
        "license_type": "hybrid",
        "vendor": "Oracle",
        "remark": "Community + Enterprise",
    },
    "MariaDB": {
        "type_code": "MARIADB",
        "name": "MariaDB",
        "category": "relational",
        "license_type": "open_source",
        "vendor": "MariaDB",
        "remark": "MySQL fork, fully open source",
    },
    "Db2": {
        "type_code": "DB2",
        "name": "Db2",
        "category": "relational",
        "license_type": "commercial",
        "vendor": "IBM",
        "remark": "IBM Db2",
    },
    "Redis": {
        "type_code": "REDIS",
        "name": "Redis",
        "category": "cache",
        "license_type": "hybrid",
        "vendor": "Redis",
        "remark": "In-memory data store",
    },
    "MongoDB": {
        "type_code": "MONGODB",
        "name": "MongoDB",
        "category": "nosql",
        "license_type": "hybrid",
        "vendor": "MongoDB",
        "remark": "Document database (SSPL)",
    },
    "Elasticsearch": {
        "type_code": "ELASTICSEARCH",
        "name": "Elasticsearch",
        "category": "search",
        "license_type": "hybrid",
        "vendor": "Elastic",
        "remark": "Search and analytics engine",
    },
    "Informix": {
        "type_code": "INFORMIX",
        "name": "Informix",
        "category": "relational",
        "license_type": "commercial",
        "vendor": "IBM",
        "remark": "IBM Informix",
    },
    "InformixAP": {
        "type_code": "INFORMIXAP",
        "name": "InformixAP",
        "category": "relational",
        "license_type": "commercial",
        "vendor": "IBM",
        "remark": "IBM Informix Advanced",
    },
}

def _normalize_db_type_key(value: Any) -> str:
    text = str(value).strip() if value not in (None, "") else ""
    return re.sub(r"[\s_]+", "", text).lower()


DB_TYPE_ALIASES = {
    _normalize_db_type_key("Oracle"): "Oracle",
    _normalize_db_type_key("PostgreSQL"): "PostgreSQL",
    _normalize_db_type_key("SQL Server"): "SQL Server",
    _normalize_db_type_key("MySQL"): "MySQL",
    _normalize_db_type_key("MariaDB"): "MariaDB",
    _normalize_db_type_key("Db2"): "Db2",
    _normalize_db_type_key("Redis"): "Redis",
    _normalize_db_type_key("MongoDB"): "MongoDB",
    _normalize_db_type_key("Elasticsearch"): "Elasticsearch",
    _normalize_db_type_key("Informix"): "Informix",
    _normalize_db_type_key("InformixAP"): "InformixAP",
    _normalize_db_type_key("Postgresql"): "PostgreSQL",
    _normalize_db_type_key("SQLServer"): "SQL Server",
}


def _normalize_cluster_type_key(value: Any) -> str:
    text = str(value).strip() if value not in (None, "") else ""
    return re.sub(r"[\s_\-]+", "", text).upper()

BASE_ROLE_MAP = {
    "Master": "primary",
    "Slave": "standby",
    "SINGLE": "single",
}

ENGINE_ROLE_MAP = {
    ("Oracle", "Master"): "primary",
    ("Oracle", "Slave"): "physical_standby",
    ("Oracle", "SINGLE"): "single",
    ("PostgreSQL", "Master"): "primary",
    ("PostgreSQL", "Slave"): "standby",
    ("PostgreSQL", "SINGLE"): "single",
    ("SQL Server", "Master"): "primary",
    ("SQL Server", "Slave"): "secondary",
    ("SQL Server", "SINGLE"): "single",
    ("MySQL", "Master"): "primary",
    ("MySQL", "Slave"): "replica",
    ("MySQL", "SINGLE"): "single",
    ("MariaDB", "Master"): "primary",
    ("MariaDB", "Slave"): "replica",
    ("MariaDB", "SINGLE"): "single",
    ("Redis", "Master"): "master",
    ("Redis", "Slave"): "replica",
    ("Redis", "SINGLE"): "single",
    ("MongoDB", "Master"): "primary",
    ("MongoDB", "Slave"): "secondary",
    ("MongoDB", "SINGLE"): "single",
    ("Db2", "Master"): "primary",
    ("Db2", "Slave"): "standby",
    ("Db2", "SINGLE"): "single",
    ("Elasticsearch", "Master"): "primary",
    ("Elasticsearch", "Slave"): "replica",
    ("Elasticsearch", "SINGLE"): "single",
    ("Informix", "Master"): "primary",
    ("Informix", "Slave"): "standby",
    ("Informix", "SINGLE"): "single",
    ("InformixAP", "Master"): "primary",
    ("InformixAP", "Slave"): "standby",
    ("InformixAP", "SINGLE"): "single",
}

CLUSTER_TYPE_CATALOG = {
    "SINGLE": {
        "name": "单实例",
        "db_types": {"ORACLE", "POSTGRESQL", "MYSQL", "MARIADB", "SQLSERVER", "DB2", "REDIS", "MONGODB", "INFORMIX", "INFORMIXAP", "ELASTICSEARCH"},
        "ha_enabled": False,
    },
    "DATAGUARD": {
        "name": "Oracle Data Guard",
        "db_types": {"ORACLE"},
        "ha_enabled": True,
    },
    "RAC": {
        "name": "Oracle RAC",
        "db_types": {"ORACLE"},
        "ha_enabled": True,
    },
    "HDR": {
        "name": "Informix HDR",
        "db_types": {"INFORMIX", "INFORMIXAP"},
        "ha_enabled": True,
    },
    "PATRONI": {
        "name": "PostgreSQL Patroni",
        "db_types": {"POSTGRESQL"},
        "ha_enabled": True,
    },
    "STREAMING_REPL": {
        "name": "PostgreSQL 流复制",
        "db_types": {"POSTGRESQL"},
        "ha_enabled": True,
    },
    "REPMGR": {
        "name": "PostgreSQL repmgr",
        "db_types": {"POSTGRESQL"},
        "ha_enabled": True,
    },
    "MGR": {
        "name": "MySQL Group Replication",
        "db_types": {"MYSQL", "MARIADB"},
        "ha_enabled": True,
    },
    "REDIS_SENTINEL": {
        "name": "Redis Sentinel",
        "db_types": {"REDIS"},
        "ha_enabled": True,
    },
    "REDIS_CLUSTER": {
        "name": "Redis Cluster",
        "db_types": {"REDIS"},
        "ha_enabled": True,
    },
    "MONGODB_RS": {
        "name": "MongoDB Replica Set",
        "db_types": {"MONGODB"},
        "ha_enabled": True,
    },
    "MONGODB_SHARD": {
        "name": "MongoDB Sharding Cluster",
        "db_types": {"MONGODB"},
        "ha_enabled": True,
    },
}

CLUSTER_TYPE_ALIASES = {
    _normalize_cluster_type_key("SINGLE"): "SINGLE",
    _normalize_cluster_type_key("single"): "SINGLE",
    _normalize_cluster_type_key("DATAGUARD"): "DATAGUARD",
    _normalize_cluster_type_key("Data Guard"): "DATAGUARD",
    _normalize_cluster_type_key("DG"): "DATAGUARD",
    _normalize_cluster_type_key("RAC"): "RAC",
    _normalize_cluster_type_key("HDR"): "HDR",
    _normalize_cluster_type_key("Informix HDR"): "HDR",
    _normalize_cluster_type_key("PATRONI"): "PATRONI",
    _normalize_cluster_type_key("STREAMING_REPL"): "STREAMING_REPL",
    _normalize_cluster_type_key("STREAMINGREPL"): "STREAMING_REPL",
    _normalize_cluster_type_key("PG_STREAMING"): "STREAMING_REPL",
    _normalize_cluster_type_key("REPMGR"): "REPMGR",
    _normalize_cluster_type_key("MGR"): "MGR",
    _normalize_cluster_type_key("REDIS_SENTINEL"): "REDIS_SENTINEL",
    _normalize_cluster_type_key("REDIS_CLUSTER"): "REDIS_CLUSTER",
    _normalize_cluster_type_key("MONGODB_RS"): "MONGODB_RS",
    _normalize_cluster_type_key("MONGODB_SHARD"): "MONGODB_SHARD",
}

SUPPORTED_CLUSTER_TYPES = set(CLUSTER_TYPE_CATALOG)

REQUIRED_HEADERS = [
    "CLUSTER_TYPE",
    "NODE_ROLE",
    "系統名稱",
    "IP",
    "實例名稱",
    "PORT",
    "DB類型",
    "國家",
    "部署類型",
    "資源提供方",
    "廠區",
    "機房位置",
]


def normalize_db_type(raw: str) -> str:
    key = _normalize_db_type_key(raw)
    if not key or key not in DB_TYPE_ALIASES:
        raise ValueError(f"Unsupported DB type: {_normalize_text(raw)}")
    return DB_TYPE_ALIASES[key]


def maybe_normalize_db_type(raw: str) -> Optional[str]:
    key = _normalize_db_type_key(raw)
    if not key:
        return None
    return DB_TYPE_ALIASES.get(key)


def normalize_cluster_type(raw: str) -> str:
    key = _normalize_cluster_type_key(raw)
    if not key or key not in CLUSTER_TYPE_ALIASES:
        raise ValueError(f"Unsupported cluster type: {_normalize_text(raw)}")
    return CLUSTER_TYPE_ALIASES[key]


def maybe_normalize_cluster_type(raw: str) -> Optional[str]:
    key = _normalize_cluster_type_key(raw)
    if not key:
        return None
    return CLUSTER_TYPE_ALIASES.get(key)


def _classify_import_issue(message: str) -> str:
    text = _normalize_text(message)
    if "缺少表头" in text:
        return "schema"
    if "不能为空" in text or "格式不正确" in text or "不支持" in text or "不适配" in text:
        return "validation"
    if "冲突" in text or "重复" in text or "已存在" in text:
        return "conflict"
    return "other"


def _build_issue_groups(messages: list[str]) -> list[dict[str, Any]]:
    labels = {
        "schema": "表结构问题",
        "validation": "校验问题",
        "conflict": "冲突 / 重复",
        "other": "其他问题",
    }
    buckets: dict[str, list[str]] = {key: [] for key in labels}
    for message in messages:
        buckets[_classify_import_issue(message)].append(message)
    return [
        {
            "key": key,
            "label": labels[key],
            "count": len(items),
            "items": items,
        }
        for key, items in buckets.items()
        if items
    ]


def normalize_node_role(db_type: str, raw_role: str) -> tuple[str, str]:
    source = (raw_role or "").strip()
    base_role = BASE_ROLE_MAP.get(source, "unknown")
    engine_role = ENGINE_ROLE_MAP.get((db_type, source))
    if engine_role is None and source.upper() == "SINGLE":
        engine_role = "single"
    if engine_role is None:
        engine_role = "unknown"
    return base_role, engine_role


def normalize_db_version(raw_version: str, raw_patch: str) -> dict[str, Optional[str]]:
    version_text = _normalize_text(raw_version)
    patch_text = _normalize_text(raw_patch) or None
    if not version_text:
        return {
            "version_code": None,
            "version_name": None,
            "patch_version": patch_text,
            "architecture_bits": None,
        }

    matches = re.findall(r"[\(（]([^()（）]+)[\)）]", version_text)
    architecture_bits = None
    for chunk in matches:
        chunk_text = _normalize_text(chunk)
        if re.search(r"\b(32|64)\s*bit\b", chunk_text, re.IGNORECASE):
            architecture_bits = re.sub(r"\s+", "", chunk_text)
            break

    version_name = re.sub(r"[\(（][^()（）]+[\)）]", "", version_text)
    version_name = re.sub(r"\s+", " ", version_name).strip()
    code_match = re.search(r"\d+(?:\.\d+)+", version_name)
    version_code = code_match.group(0) if code_match else version_name
    return {
        "version_code": version_code,
        "version_name": version_name,
        "patch_version": patch_text,
        "architecture_bits": architecture_bits,
    }


def build_cluster_code(
    system_name: str,
    db_type: str,
    cluster_type: str,
    port: Optional[str | int],
) -> str:
    seed = f"{system_name}|{db_type}|{cluster_type}|{port or ''}"
    digest = hashlib.sha1(seed.encode("utf-8")).hexdigest()[:10].upper()
    db_token = db_type.upper().replace(" ", "")
    cluster_token = (cluster_type or "UNKNOWN").upper()
    return f"CLU-{db_token}-{cluster_token}-{digest}"


def _normalize_text(value: Any) -> str:
    return str(value).strip() if value not in (None, "") else ""


def _make_code(prefix: str, seed: str) -> str:
    digest = hashlib.sha1(seed.encode("utf-8")).hexdigest()[:12].upper()
    return f"{prefix}-{digest}"


class DbopsImportService:
    CONTACT_ROLE_IMPORT_MAP = [
        ("BUSINESS_MANAGER", "business_manager", "business_manager_contact"),
        ("DBA_OWNER", "dba_owner", None),
    ]

    COLUMN_MAP = {
        "CLUSETR_NO": "source_cluster_no",
        "CLUSTER_TYPE": "cluster_type",
        "NODE_ROLE": "source_node_role",
        "系統名稱": "system_name",
        "IP": "ip",
        "實例名稱": "instance_name",
        "PORT": "port",
        "機器類型": "server_type",
        "DNS_NAME": "dns_name",
        "VIP": "vip",
        "主機管理員": "host_admin",
        "主機管理員聯繫方式": "host_admin_contact",
        "OS管理員": "os_admin",
        "OS管理員聯繫方式": "os_admin_contact",
        "業務負責人": "app_owner",
        "業務負責人聯繫方式": "app_owner_contact",
        "業務主管(必填)": "business_manager",
        "業務主管聯繫方式": "business_manager_contact",
        "業務歸屬主管(必填)": "business_belong_manager",
        "DBA負責人(必填)": "dba_owner",
        "業務單位": "business_unit",
        "部門": "department",
        "操作系統": "os_name",
        "OS版本": "os_version",
        "DB類型": "db_type",
        "DB版本": "db_version",
        "Patch版本": "db_patch",
        "INSTANCE_DESC": "instance_desc",
        "db_account": "db_account",
        "db_password": "db_password_raw",
        "ORACLE賬號": "os_oracle",
        "ROOT賬號": "os_root",
        "Hostname": "hostname",
        "CPU(Cores)": "cpu_cores",
        "記憶體(G)": "memory_gb",
        "磁碟空間(G)": "disk_gb",
        "數據庫空間(G)": "db_size_gb",
        "資源歸屬": "business_group",
        "國家": "country",
        "部署類型": "deploy_type",
        "資源提供方": "provider",
        "廠區": "factory_area",
        "機房位置": "room_location",
        "重要等級": "biz_level",
        "備份方式": "backup_type",
        "數據庫本地備份策略": "local_backup_policy",
        "備份管理策略": "backup_manage_policy",
        "審計": "audit_tag",
        "異地備份位置": "remote_backup_location",
    }

    @staticmethod
    def parse_excel_rows(file_stream: BytesIO) -> tuple[list[dict[str, Any]], list[str]]:
        workbook = load_workbook(file_stream, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        workbook.close()
        if len(rows) < 2:
            return [], []

        headers = [str(value).strip() if value else "" for value in rows[0]]
        items: list[dict[str, Any]] = []
        for row in rows[1:]:
            if not row or not any(value not in (None, "") for value in row):
                continue
            payload = {"_headers": headers, "_row": list(row)}
            for index, value in enumerate(row):
                if index < len(headers) and headers[index]:
                    payload[headers[index]] = value
            items.append(payload)
        return items, headers

    @staticmethod
    def validate_headers(headers: list[str]) -> list[str]:
        missing = [header for header in REQUIRED_HEADERS if header not in headers]
        return [f"缺少表头: {header}" for header in missing]

    @staticmethod
    def _validate_ip(ip: str) -> bool:
        return bool(re.match(r"^(\d{1,3}\.){3}\d{1,3}$", ip))

    @classmethod
    def map_excel_row_to_fields(cls, row: dict[str, Any]) -> dict[str, Any]:
        fields: dict[str, Any] = {}
        for excel_key, field_name in cls.COLUMN_MAP.items():
            if excel_key in row and row[excel_key] not in (None, ""):
                fields[field_name] = row[excel_key]
        db_type = normalize_db_type(_normalize_text(fields.get("db_type")))
        base_role, engine_role = normalize_node_role(db_type, _normalize_text(fields.get("source_node_role")))
        instance_name = _normalize_text(fields.get("instance_name"))
        cluster_type_raw = _normalize_text(fields.get("cluster_type"))
        cluster_type = maybe_normalize_cluster_type(cluster_type_raw) or cluster_type_raw.upper()
        fields["db_type"] = db_type
        fields["node_role"] = base_role
        fields["engine_role"] = engine_role
        fields["source_node_role"] = _normalize_text(fields.get("source_node_role"))
        fields["instance_name"] = instance_name
        fields["cluster_type"] = cluster_type
        fields["cluster_type_raw"] = cluster_type_raw
        fields.update(normalize_db_version(fields.get("db_version"), fields.get("db_patch")))
        fields["cluster_code"] = build_cluster_code(
            system_name=_normalize_text(fields.get("system_name")),
            db_type=db_type,
            cluster_type=cluster_type,
            port=fields.get("port"),
        )
        return fields

    @staticmethod
    def map_excel_row_to_staging_payload(row_no: int, row: dict[str, Any], source_file_name: str) -> dict[str, Any]:
        normalized_db_type = maybe_normalize_db_type(row.get("DB類型"))
        normalized_cluster_type = maybe_normalize_cluster_type(row.get("CLUSTER_TYPE"))
        normalized_node_role = BASE_ROLE_MAP.get(_normalize_text(row.get("NODE_ROLE")), "unknown")
        normalized_engine_role = None
        if normalized_db_type:
            normalized_engine_role = normalize_node_role(
                normalized_db_type,
                _normalize_text(row.get("NODE_ROLE")),
            )[1]
        return {
            "source_file_name": source_file_name,
            "row_no": row_no,
            "system_name": row.get("系統名稱"),
            "source_cluster_no": row.get("CLUSETR_NO"),
            "cluster_type": normalized_cluster_type or _normalize_text(row.get("CLUSTER_TYPE")) or None,
            "node_role": row.get("NODE_ROLE"),
            "normalized_node_role": normalized_node_role,
            "normalized_engine_role": normalized_engine_role,
            "instance_name": row.get("實例名稱"),
            "port": row.get("PORT"),
            "db_type": row.get("DB類型"),
            "normalized_db_type": normalized_db_type,
            "db_version": row.get("DB版本"),
            "db_patch": row.get("Patch版本"),
            "os_name": row.get("操作系統"),
            "os_version": row.get("OS版本"),
            "ip": row.get("IP"),
            "hostname": row.get("Hostname"),
            "server_type": row.get("機器類型"),
            "cpu_cores": row.get("CPU(Cores)"),
            "memory_gb": row.get("記憶體(G)"),
            "disk_gb": row.get("磁碟空間(G)"),
            "business_group": row.get("資源歸屬"),
            "country": row.get("國家"),
            "deploy_type": row.get("部署類型"),
            "provider": row.get("資源提供方"),
            "factory_area": row.get("廠區"),
            "room_location": row.get("機房位置"),
            "dns_name": row.get("DNS_NAME"),
            "vip": row.get("VIP"),
            "host_admin": row.get("主機管理員"),
            "host_admin_contact": row.get("主機管理員聯繫方式"),
            "os_admin": row.get("OS管理員"),
            "os_admin_contact": row.get("OS管理員聯繫方式"),
            "app_owner": row.get("業務負責人"),
            "app_owner_contact": row.get("業務負責人聯繫方式"),
            "business_manager": row.get("業務主管(必填)"),
            "business_manager_contact": row.get("業務主管聯繫方式"),
            "business_belong_manager": row.get("業務歸屬主管(必填)"),
            "dba_owner": row.get("DBA負責人(必填)"),
            "business_unit": row.get("業務單位"),
            "department": row.get("部門"),
            "biz_level": row.get("重要等級"),
            "backup_type": row.get("備份方式"),
            "local_backup_policy": row.get("數據庫本地備份策略"),
            "backup_manage_policy": row.get("備份管理策略"),
            "audit_tag": row.get("審計"),
            "remote_backup_location": row.get("異地備份位置"),
            "db_account": row.get("db_account"),
            "db_password_raw": row.get("db_password"),
            "os_oracle": row.get("ORACLE賬號"),
            "os_root": row.get("ROOT賬號"),
            "raw_payload": {
                key: value
                for key, value in row.items()
                if not key.startswith("_")
            },
        }

    @classmethod
    def validate_row(cls, row_num: int, fields: dict[str, Any]) -> list[str]:
        errors: list[str] = []
        if not _normalize_text(fields.get("system_name")):
            errors.append(f"行{row_num}: 系統名稱不能为空")
        ip = _normalize_text(fields.get("ip"))
        if not ip:
            errors.append(f"行{row_num}: IP 不能为空")
        elif not cls._validate_ip(ip):
            errors.append(f"行{row_num}: IP 格式不正确 ({ip})")
        if not _normalize_text(fields.get("instance_name")):
            errors.append(f"行{row_num}: 實例名稱不能为空")
        cluster_type_raw = _normalize_text(fields.get("cluster_type_raw")) or _normalize_text(fields.get("cluster_type"))
        if not cluster_type_raw:
            errors.append(f"行{row_num}: CLUSTER_TYPE 不能为空")
        else:
            cluster_type = _normalize_text(fields.get("cluster_type")).upper() or cluster_type_raw.upper()
            if cluster_type not in SUPPORTED_CLUSTER_TYPES:
                errors.append(
                    f"行{row_num}: CLUSTER_TYPE 不支持 ({cluster_type})，当前仅支持 "
                    f"{', '.join(sorted(SUPPORTED_CLUSTER_TYPES))}"
                )
            else:
                db_type_name = maybe_normalize_db_type(fields.get("db_type")) or _normalize_text(fields.get("db_type"))
                db_type_code = DB_TYPE_CATALOG.get(db_type_name, {}).get("type_code", _normalize_text(fields.get("db_type")).upper())
                allowed_db_types = CLUSTER_TYPE_CATALOG.get(cluster_type, {}).get("db_types", set())
                if db_type_code and allowed_db_types and db_type_code not in allowed_db_types:
                    errors.append(
                        f"行{row_num}: CLUSTER_TYPE {cluster_type} 不适配 DB類型 {db_type_name}，当前仅支持 "
                        f"{', '.join(sorted(allowed_db_types))}"
                    )
        port = _normalize_text(fields.get("port"))
        if port and not port.isdigit():
            errors.append(f"行{row_num}: PORT 必须是数字 ({port})")
        cpu_cores = _normalize_text(fields.get("cpu_cores"))
        if cpu_cores and not cpu_cores.isdigit():
            errors.append(f"行{row_num}: CPU(Cores) 必须是数字 ({cpu_cores})")
        return errors

    @staticmethod
    def _existing_record_messages(db: Optional[Session], fields: dict[str, Any]) -> list[str]:
        if db is None:
            return []
        messages: list[str] = []
        system_name = _normalize_text(fields.get("system_name"))
        ip = _normalize_text(fields.get("ip"))
        cluster_code = _normalize_text(fields.get("cluster_code"))
        instance_name = _normalize_text(fields.get("instance_name"))
        port = _normalize_text(fields.get("port"))
        if ip and db.query(Server).filter_by(ip_address=ip).first():
            messages.append(f"覆盖风险: 主机 IP {ip} 已存在，导入会更新现有主机记录")
        if system_name and db.query(BusinessSystem).filter_by(system_name=system_name).first():
            messages.append(f"覆盖风险: 系统名称 {system_name} 已存在，导入会更新现有业务系统")
        if cluster_code and db.query(Cluster).filter_by(cluster_code=cluster_code).first():
            messages.append(f"覆盖风险: 集群 {cluster_code} 已存在，导入会更新现有集群扩展字段")
        if cluster_code and instance_name and db.query(DbInstance).filter_by(
            cluster_id=db.query(Cluster).filter_by(cluster_code=cluster_code).first().id if db.query(Cluster).filter_by(cluster_code=cluster_code).first() else None,
            instance_name=instance_name,
            port=int(port) if port.isdigit() else None,
            server_id=db.query(Server).filter_by(ip_address=ip).first().id if ip and db.query(Server).filter_by(ip_address=ip).first() else None,
        ).first():
            messages.append(f"覆盖风险: 实例 {instance_name} / {port or '-'} 已存在，导入会更新现有实例")
        return messages

    @classmethod
    def preview_import(
        cls,
        file_stream: BytesIO,
        source_file_name: str,
        db: Optional[Session] = None,
        import_mode: str = "新增",
    ) -> dict[str, Any]:
        file_stream.seek(0)
        rows, headers = cls.parse_excel_rows(file_stream)
        header_errors = cls.validate_headers(headers)
        if header_errors:
            issue_groups = _build_issue_groups(header_errors)
            return {
                "total": 0,
                "success_count": 0,
                "error_count": len(header_errors),
                "conflict_count": sum(group["count"] for group in issue_groups if group["key"] == "conflict"),
                "errors": header_errors,
                "issue_groups": issue_groups,
                "import_mode": import_mode,
                "stage": "failed",
                "stage_label": "表头校验失败",
                "progress": 100,
                "items": [],
                "debug_headers": headers,
            }

        items: list[dict[str, Any]] = []
        errors: list[str] = []
        warnings: list[str] = []
        success_count = 0
        for row_num, row in enumerate(rows, start=2):
            try:
                fields = cls.map_excel_row_to_fields(row)
                row_errors = cls.validate_row(row_num, fields)
                row_warnings = cls._existing_record_messages(db, fields)
            except ValueError as exc:
                row_errors = [f"行{row_num}: {exc}"]
                fields = {}
                row_warnings = []

            status = "error" if row_errors else "ok"
            if row_errors:
                errors.extend(row_errors)
            else:
                success_count += 1
            warnings.extend(row_warnings)

            items.append({
                "row_num": row_num,
                "status": status,
                "errors": row_errors,
                "warnings": row_warnings,
                "fields": fields,
                "source_file_name": source_file_name,
                "raw_row": row,
            })

        issue_groups = _build_issue_groups(errors)
        conflict_count = sum(group["count"] for group in issue_groups if group["key"] == "conflict")
        return {
            "total": len(rows),
            "success_count": success_count,
            "error_count": len(errors),
            "warning_count": len(warnings),
            "conflict_count": conflict_count,
            "errors": errors,
            "warnings": warnings,
            "issue_groups": issue_groups,
            "import_mode": import_mode,
            "stage": "previewed",
            "stage_label": "预览完成",
            "progress": 100,
            "items": items,
            "debug_headers": headers,
        }

    @staticmethod
    def upsert_site(db: Session, fields: dict[str, Any]) -> Site:
        site_code_seed = "|".join([
            _normalize_text(fields.get("country")),
            _normalize_text(fields.get("deploy_type")),
            _normalize_text(fields.get("provider")),
            _normalize_text(fields.get("factory_area")),
            _normalize_text(fields.get("room_location")),
        ])
        site_code = _make_code("SITE", site_code_seed)
        site = db.query(Site).filter_by(site_code=site_code).first()
        if not site:
            site = Site(
                site_code=site_code,
                country=_normalize_text(fields.get("country")),
                deploy_type=_normalize_text(fields.get("deploy_type")),
                provider=_normalize_text(fields.get("provider")),
                factory_area=_normalize_text(fields.get("factory_area")),
                room_location=_normalize_text(fields.get("room_location")),
            )
            db.add(site)
            db.flush()
        return site

    @staticmethod
    def upsert_os_version(db: Session, fields: dict[str, Any]) -> OsVersion:
        os_name = _normalize_text(fields.get("os_name"))
        version_name = _normalize_text(fields.get("os_version"))
        os_code = _make_code("OS", f"{os_name}|{version_name}")
        os_version = db.query(OsVersion).filter_by(os_code=os_code).first()
        if not os_version:
            os_version = OsVersion(
                os_code=os_code,
                os_name=os_name or "Unknown",
                version_name=version_name or "Unknown",
            )
            db.add(os_version)
            db.flush()
        return os_version

    @staticmethod
    def upsert_db_type(db: Session, fields: dict[str, Any]) -> DbType:
        db_type_name = normalize_db_type(_normalize_text(fields.get("db_type")))
        catalog = DB_TYPE_CATALOG[db_type_name]
        db_type = db.query(DbType).filter_by(type_code=catalog["type_code"]).first()
        if not db_type:
            db_type = DbType(**catalog)
            db.add(db_type)
            db.flush()
        else:
            db_type.name = catalog["name"]
            db_type.category = catalog["category"]
            db_type.license_type = catalog["license_type"]
            db_type.vendor = catalog["vendor"]
            db_type.remark = catalog["remark"]
            db_type.is_active = True
        return db_type

    @staticmethod
    def upsert_db_version(db: Session, db_type_id: int, fields: dict[str, Any]) -> Optional[DbVersion]:
        version_name = _normalize_text(fields.get("version_name"))
        version_code = _normalize_text(fields.get("version_code"))
        patch_version = _normalize_text(fields.get("patch_version"))
        architecture_bits = _normalize_text(fields.get("architecture_bits"))
        if not version_name:
            return None
        db_version = db.query(DbVersion).filter_by(
            db_type_id=db_type_id,
            version_code=version_code,
            patch_version=patch_version or None,
            architecture_bits=architecture_bits or None,
        ).first()
        if not db_version:
            db_version = DbVersion(
                db_type_id=db_type_id,
                version_code=version_code,
                version_name=version_name,
                patch_version=patch_version or None,
                architecture_bits=architecture_bits or None,
            )
            db.add(db_version)
            db.flush()
        return db_version

    @staticmethod
    def upsert_business_system(db: Session, fields: dict[str, Any]) -> BusinessSystem:
        system_name = _normalize_text(fields.get("system_name"))
        business_system = db.query(BusinessSystem).filter_by(system_name=system_name).first()
        if not business_system:
            business_system = BusinessSystem(
                system_code=_make_code("SYS", system_name),
                system_name=system_name,
                business_unit=_normalize_text(fields.get("business_unit")) or None,
                department=_normalize_text(fields.get("department")) or None,
                biz_level=_normalize_text(fields.get("biz_level")) or None,
                status="active",
                extra_attrs={},
            )
            db.add(business_system)
            db.flush()
        else:
            if _normalize_text(fields.get("business_unit")):
                business_system.business_unit = _normalize_text(fields.get("business_unit"))
            if _normalize_text(fields.get("department")):
                business_system.department = _normalize_text(fields.get("department"))
            if _normalize_text(fields.get("biz_level")):
                business_system.biz_level = _normalize_text(fields.get("biz_level"))
        return business_system

    @staticmethod
    def _normalize_phone(value: Any) -> str:
        return re.sub(r"\D+", "", _normalize_text(value))

    @staticmethod
    def _normalize_email(value: Any) -> str:
        return _normalize_text(value).lower()

    @classmethod
    def resolve_contact(
        cls,
        db: Session,
        name: Any,
        phone: Any = None,
        email: Any = None,
        employee_no: Any = None,
    ) -> Optional[Contact]:
        clean_name = _normalize_text(name)
        clean_phone = cls._normalize_phone(phone)
        clean_email = cls._normalize_email(email)
        clean_employee_no = _normalize_text(employee_no)
        if not clean_name:
            return None
        if clean_employee_no:
            contact = db.query(Contact).filter_by(employee_no=clean_employee_no).first()
            if contact:
                return contact
        if clean_email:
            contact = db.query(Contact).filter_by(email=clean_email).first()
            if contact:
                return contact
        if clean_phone:
            contact = db.query(Contact).filter_by(phone=clean_phone).first()
            if contact:
                return contact
        contact = db.query(Contact).filter_by(contact_name=clean_name).first()
        return contact

    @staticmethod
    def upsert_business_contact_link(db: Session, business_system_id: int, contact_id: int, role_code: str) -> None:
        pending = getattr(db, "new", ())
        for obj in pending:
            if not isinstance(obj, BusinessSystemContact):
                continue
            if (
                obj.business_system_id == business_system_id
                and obj.contact_id == contact_id
                and obj.role_code == role_code
            ):
                return
        link = db.query(BusinessSystemContact).filter_by(
            business_system_id=business_system_id,
            contact_id=contact_id,
            role_code=role_code,
        ).first()
        if not link:
            db.add(BusinessSystemContact(
                business_system_id=business_system_id,
                contact_id=contact_id,
                role_code=role_code,
            ))

    @staticmethod
    def upsert_server(db: Session, fields: dict[str, Any], site_id: int, os_version_id: Optional[int]) -> Server:
        ip_address = _normalize_text(fields.get("ip"))
        server = db.query(Server).filter_by(ip_address=ip_address).first()
        if not server:
            server_code = _make_code("SRV", ip_address)
            server = Server(
                server_code=server_code,
                hostname=_normalize_text(fields.get("hostname")) or None,
                ip_address=ip_address,
                site_id=site_id,
                os_version_id=os_version_id,
                cpu_cores=int(fields.get("cpu_cores")) if _normalize_text(fields.get("cpu_cores")) else None,
                memory_gb=fields.get("memory_gb"),
                disk_gb=fields.get("disk_gb"),
                server_type=_normalize_text(fields.get("server_type")) or None,
                business_group=_normalize_text(fields.get("business_group")) or None,
                extra_attrs={"dns_name": _normalize_text(fields.get("dns_name")) or None},
            )
            db.add(server)
            db.flush()
        else:
            server.site_id = site_id
            server.os_version_id = os_version_id
        return server

    @staticmethod
    def upsert_cluster(
        db: Session,
        fields: dict[str, Any],
        business_system_id: int,
        db_type_id: int,
    ) -> Cluster:
        cluster = db.query(Cluster).filter_by(cluster_code=fields["cluster_code"]).first()
        cluster_type = normalize_cluster_type(_normalize_text(fields.get("cluster_type")))
        extra_attrs = {
            "source_cluster_no": _normalize_text(fields.get("source_cluster_no")) or None,
            "engine_role": _normalize_text(fields.get("engine_role")) or "unknown",
            "source_node_role": _normalize_text(fields.get("source_node_role")) or None,
        }
        if not cluster:
            cluster = Cluster(
                cluster_code=fields["cluster_code"],
                cluster_name=_normalize_text(fields.get("system_name")),
                business_system_id=business_system_id,
                db_type_id=db_type_id,
                cluster_type=cluster_type,
                ha_enabled=CLUSTER_TYPE_CATALOG.get(cluster_type, {}).get("ha_enabled", cluster_type != "SINGLE"),
                extra_attrs=extra_attrs,
            )
            db.add(cluster)
            db.flush()
        else:
            cluster.extra_attrs = {**(cluster.extra_attrs or {}), **extra_attrs}
        return cluster

    @staticmethod
    def upsert_cluster_vips(db: Session, cluster_id: int, fields: dict[str, Any]) -> None:
        vip_value = _normalize_text(fields.get("vip"))
        if not vip_value:
            return
        for vip in [item.strip() for item in vip_value.split(",") if item.strip()]:
            existing = db.query(ClusterVip).filter_by(cluster_id=cluster_id, vip_address=vip).first()
            if not existing:
                db.add(ClusterVip(cluster_id=cluster_id, vip_address=vip))

    @staticmethod
    def upsert_db_instance(
        db: Session,
        fields: dict[str, Any],
        cluster_id: int,
        server_id: int,
        db_type_id: int,
        db_version_id: Optional[int],
    ) -> tuple[DbInstance, bool]:
        instance_name = _normalize_text(fields.get("instance_name"))
        port = int(fields.get("port")) if _normalize_text(fields.get("port")) else None
        instance = db.query(DbInstance).filter_by(
            cluster_id=cluster_id,
            server_id=server_id,
            instance_name=instance_name,
            port=port,
        ).first()
        extra_attrs = {
            "engine_role": _normalize_text(fields.get("engine_role")) or "unknown",
            "source_node_role": _normalize_text(fields.get("source_node_role")) or None,
            "source_cluster_no": _normalize_text(fields.get("source_cluster_no")) or None,
            "db_patch": _normalize_text(fields.get("db_patch")) or None,
            "instance_desc": _normalize_text(fields.get("instance_desc")) or None,
        }
        created = False
        if not instance:
            created = True
            instance = DbInstance(
                instance_code=_make_code("INS", f"{cluster_id}|{server_id}|{instance_name}|{port or ''}"),
                instance_name=instance_name,
                db_type_id=db_type_id,
                db_version_id=db_version_id,
                server_id=server_id,
                cluster_id=cluster_id,
                port=port,
                node_role=_normalize_text(fields.get("node_role")) or "unknown",
                db_size_gb=fields.get("db_size_gb"),
                extra_attrs=extra_attrs,
            )
            db.add(instance)
            db.flush()
        else:
            instance.server_id = server_id
            instance.db_type_id = db_type_id
            instance.db_version_id = db_version_id
            instance.node_role = _normalize_text(fields.get("node_role")) or instance.node_role
            instance.extra_attrs = {**(instance.extra_attrs or {}), **extra_attrs}
        return instance, created

    @staticmethod
    def upsert_topology_relations(db: Session, cluster_id: int) -> None:
        instances = db.query(DbInstance).filter_by(cluster_id=cluster_id).all()
        primary = next((item for item in instances if item.node_role == "primary"), None)
        standbys = [item for item in instances if item.node_role == "standby"]
        for standby in standbys:
            pending = getattr(db, "new", ())
            duplicate_pending = False
            for obj in pending:
                if not isinstance(obj, TopologyRelation):
                    continue
                if (
                    obj.cluster_id == cluster_id
                    and obj.source_instance_id == (primary.id if primary else None)
                    and obj.target_instance_id == standby.id
                    and obj.relation_type == "primary_standby"
                ):
                    duplicate_pending = True
                    break
            if duplicate_pending:
                continue
            existing = db.query(TopologyRelation).filter_by(
                cluster_id=cluster_id,
                source_instance_id=primary.id if primary else None,
                target_instance_id=standby.id,
                relation_type="primary_standby",
            ).first()
            if not existing and primary:
                db.add(TopologyRelation(
                    cluster_id=cluster_id,
                    source_instance_id=primary.id,
                    target_instance_id=standby.id,
                    relation_type="primary_standby",
                ))

    @classmethod
    def execute_import(
        cls,
        file_stream: BytesIO,
        source_file_name: str,
        uploaded_by: str,
        db: Session,
        import_mode: str = "新增",
    ) -> dict[str, Any]:
        preview = cls.preview_import(file_stream, source_file_name, db=db, import_mode=import_mode)
        success = 0
        updated = 0
        import_batch_id = str(uuid.uuid4())
        if db is None:
            return {
                "total": preview["total"],
                "success": success,
                "updated": updated,
                "error": preview["error_count"],
                "errors": preview["errors"],
                "issue_groups": preview.get("issue_groups", []),
                "conflict_count": preview.get("conflict_count", 0),
                "import_mode": import_mode,
                "stage": "completed",
                "stage_label": "导入完成",
                "progress": 100,
                "uploaded_by": uploaded_by,
                "import_batch_id": import_batch_id,
            }
        for item in preview["items"]:
            staging_payload = cls.map_excel_row_to_staging_payload(
                row_no=item["row_num"],
                row=item["raw_row"],
                source_file_name=source_file_name,
            )
            staging_payload["import_batch_id"] = import_batch_id
            staging_payload["raw_payload"]["uploaded_by"] = uploaded_by
            db.add(StagingExcelImport(**staging_payload))

            if item["status"] == "error":
                continue

            fields = item["fields"]
            site = cls.upsert_site(db, fields)
            os_version = cls.upsert_os_version(db, fields)
            db_type = cls.upsert_db_type(db, fields)
            db_version = cls.upsert_db_version(db, db_type.id, fields)
            business_system = cls.upsert_business_system(db, fields)
            server = cls.upsert_server(db, fields, site.id, os_version.id if os_version else None)
            cluster = cls.upsert_cluster(db, fields, business_system.id, db_type.id)
            cls.upsert_cluster_vips(db, cluster.id, fields)
            instance, created = cls.upsert_db_instance(
                db,
                fields,
                cluster.id,
                server.id,
                db_type.id,
                db_version.id if db_version else None,
            )
            for role_code, name_key, phone_key in cls.CONTACT_ROLE_IMPORT_MAP:
                contact = cls.resolve_contact(
                    db,
                    fields.get(name_key),
                    fields.get(phone_key) if phone_key else None,
                )
                if contact:
                    cls.upsert_business_contact_link(db, business_system.id, contact.id, role_code)
            cls.upsert_topology_relations(db, cluster.id)
            if created:
                success += 1
            else:
                updated += 1

        db.commit()
        return {
            "total": preview["total"],
            "success": success,
            "updated": updated,
            "error": preview["error_count"],
            "errors": preview["errors"],
            "issue_groups": preview.get("issue_groups", []),
            "conflict_count": preview.get("conflict_count", 0),
            "import_mode": import_mode,
            "stage": "completed",
            "stage_label": "导入完成",
            "progress": 100,
            "uploaded_by": uploaded_by,
            "import_batch_id": import_batch_id,
        }

    @staticmethod
    def list_import_batches(db: Session, uploaded_by: Optional[str] = None) -> list[dict[str, Any]]:
        rows = db.query(StagingExcelImport).all()
        batches: dict[str, dict[str, Any]] = {}
        for row in rows:
            batch_id = row.import_batch_id or "unknown"
            owner = (row.raw_payload or {}).get("uploaded_by")
            if uploaded_by and owner != uploaded_by:
                continue
            current = batches.setdefault(batch_id, {
                "import_batch_id": batch_id,
                "source_file_name": row.source_file_name,
                "uploaded_by": owner,
                "total_rows": 0,
                "imported_at": row.imported_at.isoformat() if row.imported_at else None,
            })
            current["total_rows"] += 1
        return sorted(batches.values(), key=lambda item: item["imported_at"] or "", reverse=True)
