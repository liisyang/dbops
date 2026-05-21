from io import BytesIO

from openpyxl import Workbook

from app.models.dbops_assets import (
    BusinessSystem,
    BusinessSystemContact,
    Cluster,
    ClusterVip,
    Contact,
    DbInstance,
    DbType,
    DbVersion,
    OsVersion,
    Server,
    Site,
    StagingExcelImport,
    TopologyRelation,
)
from app.services.dbops_import_service import (
    DbopsImportService,
    normalize_db_type,
    normalize_db_version,
    normalize_node_role,
    normalize_cluster_type,
)


def _build_workbook_bytes() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.append([
        "CLUSTER_TYPE",
        "NODE_ROLE",
        "系統名稱",
        "IP",
        "實例名稱",
        "PORT",
        "DB類型",
        "DB版本",
        "操作系統",
        "OS版本",
        "國家",
        "部署類型",
        "資源提供方",
        "廠區",
        "機房位置",
    ])
    ws.append([
        "DATAGUARD",
        "Master",
        "測試系統",
        "10.0.0.10",
        "ORCL1",
        1521,
        "Oracle",
        "19c",
        "Linux",
        "RHEL 8",
        "中國",
        "地端",
        "地端",
        "深圳",
        "A1",
    ])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _build_unsupported_cluster_type_workbook_bytes() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.append([
        "CLUSTER_TYPE",
        "NODE_ROLE",
        "系統名稱",
        "IP",
        "實例名稱",
        "PORT",
        "DB類型",
        "DB版本",
        "操作系統",
        "OS版本",
        "國家",
        "部署類型",
        "資源提供方",
        "廠區",
        "機房位置",
    ])
    ws.append([
        "DATAGUARDX",
        "Master",
        "測試系統",
        "10.0.0.10",
        "ORCL1",
        1521,
        "Oracle",
        "19c",
        "Linux",
        "RHEL 8",
        "中國",
        "地端",
        "地端",
        "深圳",
        "A1",
    ])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _build_incompatible_cluster_db_type_workbook_bytes() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.append([
        "CLUSTER_TYPE",
        "NODE_ROLE",
        "系統名稱",
        "IP",
        "實例名稱",
        "PORT",
        "DB類型",
        "DB版本",
        "操作系統",
        "OS版本",
        "國家",
        "部署類型",
        "資源提供方",
        "廠區",
        "機房位置",
    ])
    ws.append([
        "DATAGUARD",
        "Master",
        "測試系統",
        "10.0.0.10",
        "PG1",
        5432,
        "PostgreSQL",
        "15",
        "Linux",
        "RHEL 8",
        "中國",
        "地端",
        "地端",
        "深圳",
        "A1",
    ])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _build_invalid_port_workbook_bytes() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.append([
        "CLUSTER_TYPE",
        "NODE_ROLE",
        "系統名稱",
        "IP",
        "實例名稱",
        "PORT",
        "DB類型",
        "DB版本",
        "操作系統",
        "OS版本",
        "國家",
        "部署類型",
        "資源提供方",
        "廠區",
        "機房位置",
    ])
    ws.append([
        "DATAGUARD",
        "Master",
        "測試系統",
        "10.0.0.10",
        "ORCL1",
        "not-a-number",
        "Oracle",
        "19c",
        "Linux",
        "RHEL 8",
        "中國",
        "地端",
        "地端",
        "深圳",
        "A1",
    ])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _build_duplicate_contact_workbook_bytes() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.append([
        "CLUSTER_TYPE",
        "NODE_ROLE",
        "系統名稱",
        "IP",
        "實例名稱",
        "PORT",
        "DB類型",
        "DB版本",
        "操作系統",
        "OS版本",
        "國家",
        "部署類型",
        "資源提供方",
        "廠區",
        "機房位置",
        "主機管理員",
        "OS管理員",
        "業務負責人",
        "業務主管(必填)",
        "業務歸屬主管(必填)",
        "DBA負責人(必填)",
    ])
    ws.append([
        "DATAGUARD",
        "Master",
        "測試系統",
        "10.0.0.10",
        "ORCL1",
        1521,
        "Oracle",
        "19c",
        "Linux",
        "RHEL 8",
        "中國",
        "地端",
        "地端",
        "深圳",
        "A1",
        "HostA",
        "HostA",
        "AppA",
        "BizA",
        "BelongA",
        "DbaA",
    ])
    ws.append([
        "DATAGUARD",
        "Slave",
        "測試系統",
        "10.0.0.11",
        "ORCL2",
        1521,
        "Oracle",
        "19c",
        "Linux",
        "RHEL 8",
        "中國",
        "地端",
        "地端",
        "深圳",
        "A1",
        "HostA",
        "HostA",
        "AppA",
        "BizA",
        "BelongA",
        "DbaA",
    ])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _build_same_instance_name_multi_server_workbook_bytes() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.append([
        "CLUSTER_TYPE",
        "NODE_ROLE",
        "系統名稱",
        "IP",
        "實例名稱",
        "PORT",
        "DB類型",
        "DB版本",
        "操作系統",
        "OS版本",
        "國家",
        "部署類型",
        "資源提供方",
        "廠區",
        "機房位置",
    ])
    ws.append([
        "DATAGUARD",
        "Master",
        "同名實例多節點",
        "10.0.0.21",
        "CHR",
        1526,
        "Oracle",
        "Oracle 19.3.0.0.0（64bit）",
        "Linux",
        "RHEL 8",
        "中國",
        "地端",
        "地端",
        "深圳",
        "A1",
    ])
    ws.append([
        "DATAGUARD",
        "Slave",
        "同名實例多節點",
        "10.0.0.22",
        "CHR",
        1526,
        "Oracle",
        "Oracle 19.3.0.0.0 (64bit)",
        "Linux",
        "RHEL 8",
        "中國",
        "地端",
        "地端",
        "深圳",
        "A1",
    ])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _build_cluster_type_alias_workbook_bytes(cluster_type: str, db_type: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.append([
        "CLUSTER_TYPE",
        "NODE_ROLE",
        "系統名稱",
        "IP",
        "實例名稱",
        "PORT",
        "DB類型",
        "DB版本",
        "操作系統",
        "OS版本",
        "國家",
        "部署類型",
        "資源提供方",
        "廠區",
        "機房位置",
    ])
    ws.append([
        cluster_type,
        "Member",
        "别名集群",
        "10.0.0.31",
        "TEST01",
        5432,
        db_type,
        "15.4",
        "Linux",
        "RHEL 8",
        "中国",
        "地端",
        "地端",
        "深圳",
        "A1",
    ])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _build_single_sqlserver_workbook_bytes() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.append([
        "CLUSTER_TYPE",
        "NODE_ROLE",
        "系統名稱",
        "IP",
        "實例名稱",
        "PORT",
        "DB類型",
        "DB版本",
        "操作系統",
        "OS版本",
        "國家",
        "部署類型",
        "資源提供方",
        "廠區",
        "機房位置",
    ])
    ws.append([
        "SINGLE",
        "Master",
        "SQL Server 单实例",
        "10.0.0.51",
        "MSSQL01",
        1433,
        "SQL Server",
        "2019",
        "Windows",
        "Windows Server 2019",
        "中国",
        "地端",
        "地端",
        "上海",
        "A1",
    ])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _seed_contact_master_rows():
    return [
        Contact(id=100, contact_code="CT-MASTER-01", contact_name="HostA", phone=""),
        Contact(id=101, contact_code="CT-MASTER-02", contact_name="AppA", phone=""),
        Contact(id=102, contact_code="CT-MASTER-03", contact_name="BizA", phone=""),
        Contact(id=103, contact_code="CT-MASTER-04", contact_name="BelongA", phone=""),
        Contact(id=104, contact_code="CT-MASTER-05", contact_name="DbaA", phone=""),
    ]


def test_normalize_db_type_maps_known_values():
    assert normalize_db_type("Oracle") == "Oracle"
    assert normalize_db_type("Postgresql") == "PostgreSQL"
    assert normalize_db_type("SQL Server") == "SQL Server"
    assert normalize_db_type("Db2") == "Db2"
    assert normalize_db_type("MariaDB") == "MariaDB"
    assert normalize_db_type("Elasticsearch") == "Elasticsearch"
    assert normalize_db_type("Informix") == "Informix"
    assert normalize_db_type("InformixAP") == "InformixAP"


def test_normalize_cluster_type_maps_known_values():
    assert normalize_cluster_type("Redis Cluster") == "REDIS_CLUSTER"
    assert normalize_cluster_type("PG Streaming") == "STREAMING_REPL"
    assert normalize_cluster_type("Informix HDR") == "HDR"


def test_normalize_node_role_maps_excel_values():
    base_role, engine_role = normalize_node_role("Oracle", "Master")
    assert base_role == "primary"
    assert engine_role == "primary"

    base_role, engine_role = normalize_node_role("PostgreSQL", "SINGLE")
    assert base_role == "single"
    assert engine_role == "single"


def test_normalize_db_version_strips_brackets_and_extracts_bits():
    result = normalize_db_version("Oracle 11.2.0.4.0（64bit）", "")
    assert result["version_code"] == "11.2.0.4.0"
    assert result["version_name"] == "Oracle 11.2.0.4.0"
    assert result["architecture_bits"] == "64bit"


def test_preview_import_returns_standardized_item():
    result = DbopsImportService.preview_import(_build_workbook_bytes(), "hr_list.xlsx")
    assert result["total"] == 1
    assert result["error_count"] == 0
    assert result["import_mode"] == "新增"
    assert result["stage"] == "previewed"
    assert result["stage_label"] == "预览完成"
    assert result["progress"] == 100
    assert result["items"][0]["fields"]["db_type"] == "Oracle"
    assert result["items"][0]["fields"]["node_role"] == "primary"
    assert result["items"][0]["fields"]["cluster_type"] == "DATAGUARD"


def test_preview_import_rejects_unsupported_cluster_type():
    result = DbopsImportService.preview_import(_build_unsupported_cluster_type_workbook_bytes(), "hr_list.xlsx")

    assert result["total"] == 1
    assert result["success_count"] == 0
    assert result["error_count"] == 1
    assert "CLUSTER_TYPE 不支持" in result["errors"][0]
    assert result["issue_groups"][0]["label"] == "校验问题"


def test_preview_import_rejects_incompatible_cluster_and_db_type():
    result = DbopsImportService.preview_import(_build_incompatible_cluster_db_type_workbook_bytes(), "hr_list.xlsx")

    assert result["total"] == 1
    assert result["success_count"] == 0
    assert result["error_count"] == 1
    assert "不适配 DB類型" in result["errors"][0]
    assert result["issue_groups"][0]["label"] == "校验问题"


def test_preview_import_rejects_invalid_port():
    result = DbopsImportService.preview_import(_build_invalid_port_workbook_bytes(), "hr_list.xlsx")

    assert result["total"] == 1
    assert result["success_count"] == 0
    assert result["error_count"] == 1
    assert "PORT 必须是数字" in result["errors"][0]


def test_preview_import_accepts_single_sqlserver():
    result = DbopsImportService.preview_import(_build_single_sqlserver_workbook_bytes(), "hr_list.xlsx")

    assert result["total"] == 1
    assert result["success_count"] == 1
    assert result["error_count"] == 0
    assert result["items"][0]["status"] == "ok"


def test_preview_import_flags_overwrite_risk_for_existing_records():
    db = _FakeSession()
    db.seed(
        Site(site_code="SITE_1", country="中國", deploy_type="地端", provider="地端", factory_area="深圳", room_location="A1"),
        OsVersion(os_code="OS_1", os_name="Linux", version_name="RHEL 8"),
        BusinessSystem(system_code="SYS_1", system_name="測試系統"),
        DbType(type_code="ORACLE", name="Oracle", category="relational", license_type="commercial", vendor="Oracle", remark="Oracle Database"),
        Cluster(
            cluster_code="測試系統_ORACLE_DATAGUARD_1521",
            cluster_name="測試系統",
            business_system_id=3,
            db_type_id=4,
            cluster_type="DATAGUARD",
            ha_enabled=True,
        ),
        Server(server_code="SRV_1", ip_address="10.0.0.10", site_id=1, os_version_id=2),
        DbInstance(
            instance_code="INS_1",
            instance_name="ORCL1",
            db_type_id=4,
            db_version_id=None,
            server_id=6,
            cluster_id=5,
            port=1521,
            node_role="primary",
        ),
    )

    result = DbopsImportService.preview_import(_build_workbook_bytes(), "hr_list.xlsx", db=db)

    assert result["total"] == 1
    assert result["success_count"] == 1
    assert result["error_count"] == 0
    assert result["warning_count"] >= 1
    assert any("覆盖风险" in message for message in result["warnings"])
    assert result["items"][0]["warnings"]


def test_upsert_business_system_defaults_imported_rows_to_active():
    db = _FakeSession()

    business_system = DbopsImportService.upsert_business_system(
        db,
        {
            "system_name": "测试系统",
            "business_unit": "事业群",
            "department": "平台部",
            "biz_level": "重要",
        },
    )

    assert business_system.status == "active"


class _FakeQuery:
    def __init__(self, session, model):
        self.session = session
        self.model = model
        self.filters = {}

    def filter_by(self, **kwargs):
        self.filters.update(kwargs)
        return self

    def _rows(self):
        rows = self.session.store.get(self.model, [])
        for key, value in self.filters.items():
            rows = [row for row in rows if getattr(row, key) == value]
        return rows

    def first(self):
        rows = self._rows()
        return rows[0] if rows else None

    def all(self):
        return list(self._rows())

    def count(self):
        return len(self._rows())


class _FakeSession:
    def __init__(self):
        self.store = {}
        self._next_id = 1
        self.new = []

    def seed(self, *objects):
        for obj in objects:
            bucket = self.store.setdefault(type(obj), [])
            bucket.append(obj)
            if getattr(obj, "id", None) is None:
                obj.id = self._next_id
                self._next_id += 1

    def add(self, obj):
        bucket = self.store.setdefault(type(obj), [])
        if obj not in bucket:
            bucket.append(obj)
        if obj not in self.new:
            self.new.append(obj)
        if getattr(obj, "id", None) is None:
            obj.id = self._next_id
            self._next_id += 1

    def flush(self):
        self.new = []
        return None

    def query(self, model):
        return _FakeQuery(self, model)

    def commit(self):
        self.new = []
        return None


def test_execute_import_upserts_business_cluster_server_and_instance():
    db = _FakeSession()
    result = DbopsImportService.execute_import(
        _build_workbook_bytes(),
        "hr_list.xlsx",
        "admin",
        db,
    )

    assert result["success"] == 1
    assert result["updated"] == 0
    assert db.query(StagingExcelImport).count() == 1
    assert db.query(BusinessSystem).count() == 1
    assert db.query(Site).count() == 1
    assert db.query(Server).count() == 1
    assert db.query(DbType).count() == 1
    db_type = db.query(DbType).first()
    assert db_type.type_code == "ORACLE"
    assert db_type.category == "relational"
    assert db_type.license_type == "commercial"
    assert db_type.vendor == "Oracle"
    assert db.query(OsVersion).count() == 1
    assert db.query(DbVersion).count() == 1
    assert db.query(Cluster).count() == 1
    assert db.query(DbInstance).count() == 1
    assert db.query(Contact).count() == 0
    assert db.query(ClusterVip).count() == 0
    assert db.query(TopologyRelation).count() == 0
    staging_row = db.query(StagingExcelImport).first()
    assert staging_row.source_cluster_no is None
    assert staging_row.normalized_db_type == "Oracle"
    assert staging_row.normalized_node_role == "primary"
    assert staging_row.normalized_engine_role == "primary"
    assert result["import_mode"] == "新增"
    assert result["stage"] == "completed"
    assert result["stage_label"] == "导入完成"
    assert result["progress"] == 100


def test_execute_import_is_idempotent_for_same_file():
    db = _FakeSession()
    first = DbopsImportService.execute_import(_build_workbook_bytes(), "hr_list.xlsx", "admin", db)
    second = DbopsImportService.execute_import(_build_workbook_bytes(), "hr_list.xlsx", "admin", db)

    assert first["success"] == 1
    assert second["success"] == 0
    assert second["updated"] == 1
    assert db.query(StagingExcelImport).count() == 2
    assert db.query(BusinessSystem).count() == 1
    assert db.query(Site).count() == 1
    assert db.query(Server).count() == 1
    assert db.query(Cluster).count() == 1
    assert db.query(DbInstance).count() == 1


def test_execute_import_deduplicates_business_contact_links_within_batch():
    db = _FakeSession()
    db.seed(*_seed_contact_master_rows())
    result = DbopsImportService.execute_import(
        _build_duplicate_contact_workbook_bytes(),
        "hr_list.xlsx",
        "admin",
        db,
    )

    assert result["success"] == 2
    assert db.query(BusinessSystem).count() == 1
    assert db.query(Contact).count() == 5
    assert db.query(DbInstance).count() == 2
    assert db.query(TopologyRelation).count() == 1
    assert db.query(BusinessSystemContact).count() == 2


def test_execute_import_reuses_existing_contact_master_data():
    db = _FakeSession()
    db.seed(*_seed_contact_master_rows())

    result = DbopsImportService.execute_import(
        _build_duplicate_contact_workbook_bytes(),
        "hr_list.xlsx",
        "admin",
        db,
    )

    assert result["success"] == 2
    assert db.query(Contact).count() == 5
    assert db.query(BusinessSystemContact).count() == 2


def test_execute_import_keeps_same_instance_name_on_different_servers():
    db = _FakeSession()
    result = DbopsImportService.execute_import(
        _build_same_instance_name_multi_server_workbook_bytes(),
        "hr_list.xlsx",
        "admin",
        db,
    )

    assert result["success"] == 2
    assert db.query(Cluster).count() == 1
    assert db.query(Server).count() == 2
    assert db.query(DbInstance).count() == 2
    assert db.query(TopologyRelation).count() == 1


def test_execute_import_accepts_cluster_type_alias_and_canonicalizes_it():
    db = _FakeSession()
    result = DbopsImportService.execute_import(
        _build_cluster_type_alias_workbook_bytes("Redis Cluster", "Redis"),
        "hr_list.xlsx",
        "admin",
        db,
    )

    assert result["success"] == 1
    cluster = db.query(Cluster).first()
    assert cluster.cluster_type == "REDIS_CLUSTER"
    assert cluster.ha_enabled is True
    staging_row = db.query(StagingExcelImport).first()
    assert staging_row.cluster_type == "REDIS_CLUSTER"
